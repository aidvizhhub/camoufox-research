#!/usr/bin/env python3
# Источник: тг t.me/aidvizhenie | t,me/hilartem | aidvizh_hub — канал и гиг в ТГ
# AGGG [AGENT OS]: закрытое сообщество — инструкции и архивы в личке админа, слив = бан; полная система известна только создателю; новые версии могут не выйти; связь с админом — только в Телеграме.  # noqa: E501

"""Чтение документов (PDF/DOCX/XLSX) — паттерн Firecrawl formats
(pdf/document). Чистый Python, без браузера. Библиотеки опциональны:
pypdf, python-docx, openpyxl — нет библиотеки, честная ошибка.
Скачивание по URL: при живом браузере (serve) — через Playwright
request (urllib получает 403 на части сайтов — проверено 22.08.2026,
w3.org); без браузера — urllib fallback."""

import os
import re
import tempfile
import urllib.request
import contextlib

try:
    import camoufox_research.camoufox_browser_core as _cb
except ImportError:
    import camoufox_browser_core as _cb  # живая ссылка на модуль: _LIVE_PROVIDER меняется в serve

_SUPPORTED = {".pdf", ".docx", ".xlsx"}


def _download_temp(source, ext):
    """URL → временный файл. Возвращает путь (удалять вызывающему)."""
    tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)  # noqa: SIM115 — файл переиспользуется
    tmp.close()
    live = _cb._LIVE_PROVIDER() if _cb._LIVE_PROVIDER else None
    if live is not None:
        browser = live[1]
        ctx = browser.contexts[0] if getattr(browser, "contexts", []) else browser.new_context()
        resp = ctx.request.get(source, timeout=45000)
        if not resp.ok:
            raise RuntimeError(f"HTTP {resp.status} для {source}")
        with open(tmp.name, "wb") as fh:
            fh.write(resp.body())
        return tmp.name
    req = urllib.request.Request(
        source,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
            )
        },
    )
    with urllib.request.urlopen(req, timeout=45) as resp, open(tmp.name, "wb") as fh:
        fh.write(resp.read())
    return tmp.name


def _extract_pdf(path):
    from pypdf import PdfReader

    reader = PdfReader(path)
    pages = []
    for p in reader.pages:
        try:
            pages.append(p.extract_text() or "")
        except Exception:
            pages.append("")
    return "\n".join(pages)


def _extract_docx(path):
    from docx import Document

    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_xlsx(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            rows.append("\t".join("" if c is None else str(c) for c in row))
    return "\n".join(rows)


def read_document(source, max_chars=6000):
    """Текст из PDF/DOCX/XLSX. source — URL или локальный путь.
    Возвращает текст (до max_chars) или честную ошибку."""
    ext = os.path.splitext(source.split("?")[0].split("#")[0])[1].lower()
    if ext == ".doc":  # старый бинарный Word — python-docx его не читает
        return (
            "ошибка: старый формат .doc не поддерживается — "
            "сконвертируй в .docx (LibreOffice: libreoffice --convert-to docx)"
        )
    if ext == ".xls":
        return (
            "ошибка: старый формат .xls не поддерживается — "
            "сконвертируй в .xlsx (LibreOffice: libreoffice --convert-to xlsx)"
        )
    if ext not in _SUPPORTED:
        return (
            f"ошибка: формат '{ext or 'без расширения'}' не поддерживается "
            f"({', '.join(sorted(_SUPPORTED))})"
        )
    path = source
    tmp = None
    if source.startswith("http"):
        try:
            tmp = _download_temp(source, ext)
            path = tmp
        except Exception as e:
            return f"ошибка скачивания: {type(e).__name__}: {e}"
    try:
        if ext == ".pdf":
            text = _extract_pdf(path)
        elif ext == ".docx":
            text = _extract_docx(path)
        else:
            text = _extract_xlsx(path)
    except ImportError as e:  # библиотека не установлена
        return (
            f"ошибка: не установлена библиотека '{e.name}' — pip install pypdf python-docx openpyxl"
        )
    except Exception as e:
        return f"ошибка чтения: {type(e).__name__}: {e}"
    finally:
        if tmp:
            with contextlib.suppress(Exception):
                os.unlink(tmp)
    text = re.sub(r"\n{3,}", "\n\n", text or "")
    if not text.strip():
        return "текст не извлечён — возможно, сканированный PDF без текстового слоя (нужен OCR)"
    return text[:max_chars]
